from openpyxl import Workbook

#创建工作簿
wb = Workbook()
#创建工作表
ws = wb.active

print('success')

min_row, min_col, max_row, max_col = map(int, input('请输入最小，最大行列:').split())

for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        val = input(f'输入{row}行， {col}列的数据:')
        ws.cell(row, col, val)

#按照列打印
for cells in ws.iter_cols():
    for cell in cells:
        print(cell.value,end=' ')
    print('\n')
#按照行打印
for cells in ws.iter_rows():
    for cell in cells:
        print(cell.value, end=' ')
    print('\n')
#保存
wb.save('test3.xlsx')

