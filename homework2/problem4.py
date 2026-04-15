from openpyxl import Workbook
#创建工作簿
wb = Workbook()
ws = wb.active

with open('test4.txt', 'r', encoding='utf-8') as f:
    line = f.readlines()

print(line)

for row, row_data in enumerate(line, start=1):
    row_data = row_data.strip().split(',')
    for col, val in enumerate(row_data, start=1):
        ws.cell(row, col, val)

for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        print(ws.cell(r, c).value, end=' ')
    print('\n')

wb.save('test4.xlsx')
